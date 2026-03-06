"""
将大 Excel 流式导入到 SQLite（openpyxl read_only 模式，控制内存）。
支持：多 sheet、每行文本/公式、图片导出到 static/images。
用法：设置环境变量 EXCEL_PATH 或修改 config.EXCEL_PATH 后运行
      python import_excel.py
"""
import json
import os
import sys
from pathlib import Path

# 确保可导入 config、database
sys.path.insert(0, str(Path(__file__).resolve().parent))

import config
from database import get_connection, init_db, clear_all_rows, insert_row, insert_image
from extract_xlsx_images import extract_images_from_xlsx

try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
except ImportError:
    print("请先安装: pip install openpyxl")
    sys.exit(1)


def _cell_value(cell, data_only: bool = True):
    """取单元格显示值。data_only=True 时取公式计算结果（与 Excel 显示一致）。"""
    if cell.value is None:
        return ""
    if isinstance(cell.value, (int, float)):
        return cell.value
    return str(cell.value)


def _export_image(img, sheet_name: str, row_idx: int, col_idx: int) -> str:
    """把 openpyxl 图片保存到 static/images，返回相对路径。"""
    config.IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    ext = "png"
    if hasattr(img, "ref") and img.ref:
        name = f"{sheet_name}_r{row_idx}_c{col_idx}_{id(img)}.{ext}"
    else:
        name = f"{sheet_name}_r{row_idx}_c{col_idx}_{id(img)}.{ext}"
    path = config.IMAGES_DIR / name
    if hasattr(img, "_data") and img._data():
        path.write_bytes(img._data())
    elif hasattr(img, "ref") and getattr(img, "_data", None):
        path.write_bytes(img._data())
    return f"/static/images/{name}"


def import_workbook(excel_path: str):
    if not os.path.isfile(excel_path):
        print(f"文件不存在: {excel_path}")
        print("请设置环境变量 EXCEL_PATH 或把 Excel 放到项目目录并命名为 data.xlsx")
        return

    init_db()
    conn = get_connection()
    try:
        clear_all_rows(conn)
        print("  已清空旧数据，本次导入将覆盖为最新内容。")
    finally:
        conn.close()

    print(f"开始流式读取: {excel_path}")

    # 先从 xlsx 内提取嵌入图片（仅支持通过“浮动图形”锚定的图片；DISPIMG/单元格内图片无法解析）
    print("  正在解析并提取嵌入图片...")
    cell_images = extract_images_from_xlsx(excel_path)
    if cell_images:
        print(f"  已提取 {len(cell_images)} 张图片与单元格对应关系")
    else:
        print("  未检测到可解析的嵌入图片（若为 DISPIMG/单元格内图片，当前仅能显示 [图片] 占位符）")

    # read_only=True, data_only=True 以便取公式计算结果（与 Excel 显示一致；需 Excel 曾保存过才有缓存值）
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    conn = get_connection()
    cursor = conn.cursor()
    total = 0
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 若尺寸不对可取消下一行注释
            # ws.reset_dimensions()
            row_index = 0
            # 收集当前 sheet 的图片（非 read_only 时可用 ws._images；read_only 下可能为空）
            sheet_images = {}
            try:
                if hasattr(ws, "_images") and ws._images:
                    for img in ws._images:
                        anchor = getattr(img, "anchor", None)
                        if anchor and hasattr(anchor, "_from"):
                            row, col = getattr(anchor._from, "row", 0), getattr(anchor._from, "col", 0)
                            key = (row, col)
                            rel_path = _export_image(img, sheet_name, row, col)
                            sheet_images[key] = rel_path
            except Exception:
                pass

            for row in ws.iter_rows():
                row_index += 1
                column_data = []
                for col_idx, cell in enumerate(row):
                    val = _cell_value(cell, data_only=True)
                    column_data.append(val)
                row_id = insert_row(cursor, sheet_name, row_index, column_data, formula_text=None)
                for col_idx in range(len(column_data)):
                    key_ws = (sheet_name, row_index - 1, col_idx)
                    if key_ws in cell_images:
                        insert_image(cursor, row_id, cell_images[key_ws], col_ref=f"col_{col_idx}")
                    key = (row_index - 1, col_idx)
                    if key in sheet_images:
                        insert_image(cursor, row_id, sheet_images[key], col_ref=f"col_{col_idx}")
                total += 1
                if config.IMPORT_MAX_ROWS and total >= config.IMPORT_MAX_ROWS:
                    print(f"  已达上限 {config.IMPORT_MAX_ROWS} 行，停止导入（用于生成小体积部署库）")
                    break
                if total % config.IMPORT_BATCH_SIZE == 0:
                    conn.commit()
                    print(f"  已导入 {total} 行...")
            print(f"  Sheet [{sheet_name}] 完成")
            if config.IMPORT_MAX_ROWS and total >= config.IMPORT_MAX_ROWS:
                break
        conn.commit()
        # 限制行数导入后执行 VACUUM，缩小 db 文件体积以便推送到 GitHub（≤100MB）
        if config.IMPORT_MAX_ROWS:
            print("  正在压缩数据库文件 (VACUUM)...")
            conn.execute("VACUUM")
    finally:
        cursor.close()
        conn.close()
        wb.close()

    print(f"导入完成，共 {total} 行。")


if __name__ == "__main__":
    import_workbook(config.EXCEL_PATH)
