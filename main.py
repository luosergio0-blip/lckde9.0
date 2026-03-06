"""
FastAPI 后端：搜索、详情、添加新内容、按关键字导出 Excel、静态资源。
"""
import io
import os
import shutil
import urllib.request
from pathlib import Path

import config
from database import (
    get_connection,
    init_db,
    get_row_by_id,
    get_sheet_header_labels,
    search_with_preview,
    get_sheets,
    get_sheet_counts,
    get_sheet_total,
    get_sheet_rows,
    insert_row,
    insert_image,
    get_row_ids_by_models,
    get_full_rows_by_ids,
    get_images_by_row_ids,
)
from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
except ImportError:
    Workbook = None
    OpenpyxlImage = None
    get_column_letter = None
    OneCellAnchor = None
    AnchorMarker = None
    XDRPositiveSize2D = None
    pixels_to_EMU = None

app = FastAPI(title="Excel 查询站", description="大 Excel 导入后查询与新增")

# 允许所有人访问：任意来源的浏览器都可请求本接口（局域网/公网他人访问不会跨域报错）
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 静态文件：前端 + 图片
STATIC_ROOT = Path(__file__).resolve().parent / "static"
STATIC_ROOT.mkdir(exist_ok=True)
(config.IMAGES_DIR).mkdir(parents=True, exist_ok=True)
app.mount("/static", StaticFiles(directory=str(STATIC_ROOT)), name="static")


def ensure_db_file():
    """
    若本地不存在 excel_data.db 且配置了 DB_DOWNLOAD_URL，则在启动前自动从远端下载完整数据库。
    适合将 excel_data.db 上传到 GitHub Release / 对象存储等位置，让 Render 实例启动时拉取。
    """
    db_path = Path(config.DB_PATH)
    if db_path.exists():
        return
    url = getattr(config, "DB_DOWNLOAD_URL", "") or os.environ.get("DB_DOWNLOAD_URL", "")
    if not url:
        # 未配置远端地址，则按空库启动（后续可自行导入）
        print("DB 文件不存在，且未设置 DB_DOWNLOAD_URL，将使用空数据库。")
        return
    db_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        print(f"正在从远端下载数据库文件: {url}")
        with urllib.request.urlopen(url) as resp, open(db_path, "wb") as f:
            shutil.copyfileobj(resp, f)
        print(f"数据库下载完成: {db_path}")
    except Exception as e:
        print(f"下载数据库失败，将使用空数据库: {e}")


@app.on_event("startup")
def startup():
    ensure_db_file()
    init_db()


# ---------- API ----------

class AddRowRequest(BaseModel):
    sheet_name: str
    column_data: list
    formula_text: str | None = None


class ExportByModelsRequest(BaseModel):
    """按关键字导出：传入关键字列表或一个字符串（多行/逗号分隔），导出匹配的整行到 Excel。"""
    models: list[str] | str = []


class ExportByIdsRequest(BaseModel):
    """按行 id 导出：传入当前搜索结果的行 id 列表，导出这些行的整行到 Excel。"""
    ids: list[int] = []


@app.get("/api/access-info")
def api_access_info(request: Request):
    """返回当前访问地址，便于分享给其他人（局域网或公网）。"""
    host = (request.headers.get("host") or "localhost:8000").strip()
    port = "8000"
    if ":" in host:
        host, port = host.rsplit(":", 1)
    if not port or port == "80":
        port = "8000"
    access_url = f"http://{host}:{port}" if port != "80" else f"http://{host}"
    is_local = host in ("localhost", "127.0.0.1")
    if is_local:
        tip = "您当前通过本机访问。其他人需用您电脑的 IP：在本机 CMD 输入 ipconfig，将「IPv4 地址」与 :8000 组成地址分享（如 http://192.168.1.100:8000）。请确保本机防火墙已放行 8000 端口。"
    else:
        tip = "同一局域网内其他人可用此地址访问；若需外网访问请在路由器做端口转发（外网 8000 → 本机 8000）。"
    return {"access_url": access_url, "tip": tip, "is_local": is_local}


@app.get("/api/sheets")
def api_sheets():
    """获取所有 sheet 名称及每表行数。"""
    conn = get_connection()
    try:
        counts = get_sheet_counts(conn)
        return {"sheets": [c[0] for c in counts], "counts": {c[0]: c[1] for c in counts}}
    finally:
        conn.close()


@app.get("/api/stats")
def api_stats():
    """查看库内总行数、各 sheet 行数（用于排查“暂无数据”）。"""
    conn = get_connection()
    try:
        counts = get_sheet_counts(conn)
        total = sum(c[1] for c in counts)
        return {"total_rows": total, "sheets": [{"name": c[0], "rows": c[1]} for c in counts]}
    finally:
        conn.close()


@app.get("/api/sheet/{sheet_name:path}")
def api_sheet(
    sheet_name: str,
    offset: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    order: str = Query("desc", description="asc 正序 / desc 表头置顶其余倒序"),
):
    """按工作表分页取行。order=desc 时表头置顶、其余行倒序显示。"""
    conn = get_connection()
    try:
        total = get_sheet_total(conn, sheet_name)
        rows = get_sheet_rows(conn, sheet_name, offset=offset, limit=limit, order=order if order in ("asc", "desc") else "desc")
        return {"sheet_name": sheet_name, "total": total, "offset": offset, "limit": limit, "order": order, "rows": rows}
    finally:
        conn.close()


@app.get("/api/search")
def api_search(q: str = Query("", min_length=0), limit: int = Query(100, ge=1, le=500)):
    """全文搜索（FTS + LIKE 备用），按关键字可查。"""
    conn = get_connection()
    try:
        if not q.strip():
            return {"results": [], "query": q}
        try:
            results = search_with_preview(conn, q, limit=limit)
            return {"results": results, "query": q}
        except Exception as e:
            return {"results": [], "query": q, "error": str(e)}
    finally:
        conn.close()


@app.get("/api/row/{row_id:int}")
def api_row(row_id: int):
    """获取单行详情（含列数据、公式、图片列表、该表第一行列名）。"""
    conn = get_connection()
    try:
        row = get_row_by_id(conn, row_id)
        if not row:
            raise HTTPException(status_code=404, detail="未找到该行")
        row["header_labels"] = get_sheet_header_labels(conn, row["sheet_name"])
        return row
    finally:
        conn.close()


@app.post("/api/row")
def api_add_row(body: AddRowRequest):
    """添加一行新内容。"""
    conn = get_connection()
    try:
        cursor = conn.cursor()
        # 当前最大 row_index（同 sheet）
        cursor.execute(
            "SELECT COALESCE(MAX(row_index), 0) + 1 FROM rows WHERE sheet_name = ?",
            (body.sheet_name,),
        )
        next_idx = cursor.fetchone()[0]
        row_id = insert_row(
            cursor,
            body.sheet_name,
            next_idx,
            body.column_data,
            body.formula_text,
        )
        conn.commit()
        return {"id": row_id, "sheet_name": body.sheet_name, "row_index": next_idx}
    finally:
        conn.close()


@app.post("/api/import")
def api_import_trigger():
    """触发导入（同步会很久，生产环境建议用后台任务）。"""
    from import_excel import import_workbook
    import_workbook(config.EXCEL_PATH)
    return {"status": "ok", "message": "导入已执行，请查看控制台输出"}


@app.post("/api/export-excel")
def api_export_excel(body: ExportByModelsRequest):
    """按关键字导出：传入多个关键字（每行或逗号分隔），返回包含匹配整行的 Excel 文件。"""
    if not Workbook:
        raise HTTPException(status_code=500, detail="请安装 openpyxl: pip install openpyxl")
    if isinstance(body.models, str):
        raw = [body.models]
    else:
        raw = list(body.models) if body.models else []
    terms = []
    for item in raw:
        s = (str(item).strip() if item is not None else "") or ""
        if s:
            for part in s.replace("\r\n", "\n").replace("，", ",").split("\n"):
                terms.extend(t.strip() for t in part.split(",") if t.strip())
    terms = [t for t in terms if t]
    if not terms:
        raise HTTPException(status_code=400, detail="请至少提供一个关键字")
    conn = get_connection()
    try:
        ids = get_row_ids_by_models(conn, terms, limit_per_term=2000)
        if not ids:
            raise HTTPException(status_code=404, detail="未找到任何匹配行，请检查关键字")
        rows = get_full_rows_by_ids(conn, ids)
        header_labels = get_sheet_header_labels(conn, rows[0]["sheet_name"]) if rows else None
        images_by_row = get_images_by_row_ids(conn, ids)
    finally:
        conn.close()
    buf = _build_export_xlsx(rows, "按关键字导出", header_labels=header_labels, images_by_row=images_by_row)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=export_by_models.xlsx"},
    )


def _resolve_image_path(
    path_val: str,
    row_id: int | None = None,
    row_index: int | None = None,
    export_row_index: int | None = None,
) -> Path | None:
    """将 DB 中的图片路径解析为本地文件路径。支持 cellimg_0, cellimg_1（导出顺序）或 cellimg_{row_id}/{row_index}。"""
    path_val = (path_val or "").strip()
    # 1) 路径不为空：先按 DB 存的路径找
    if path_val:
        fname = os.path.basename(path_val.replace("\\", "/"))
        if fname:
            full = config.IMAGES_DIR / fname
            if full.is_file():
                return full
        if path_val.startswith("static/") or (not path_val.startswith("/") and "static" in path_val):
            p = config.BASE_DIR / path_val.replace("\\", "/")
            if p.is_file():
                return p
        p = Path(path_val)
        if p.is_file():
            return p
    # 2) 按约定 cellimg_X 查找：优先导出顺序（cellimg_0, cellimg_1...），再 row_id / row_index
    if not config.IMAGES_DIR.is_dir():
        return None
    for ext in ("jpeg", "png", "jpg"):
        if export_row_index is not None:
            candidate = config.IMAGES_DIR / f"cellimg_{export_row_index}.{ext}"
            if candidate.is_file():
                return candidate
        if row_id is not None:
            candidate = config.IMAGES_DIR / f"cellimg_{row_id}.{ext}"
            if candidate.is_file():
                return candidate
        if row_index is not None and row_index != row_id:
            candidate = config.IMAGES_DIR / f"cellimg_{row_index}.{ext}"
            if candidate.is_file():
                return candidate
    return None


# 导出时图片在单元格内的显示尺寸（像素），用单格锚点+固定范围实现「嵌入单元格」
_EXPORT_IMAGE_CELL_SIZE = 80


def _scale_image_to_cell(img, max_wh: int = _EXPORT_IMAGE_CELL_SIZE) -> None:
    """将图片等比例缩小到不超过 max_wh，使嵌入单元格后不溢出。"""
    w = getattr(img, "width", None) or 0
    h = getattr(img, "height", None) or 0
    if w <= 0 or h <= 0:
        img.width = max_wh
        img.height = max_wh
        return
    if w <= max_wh and h <= max_wh:
        return
    if w > h:
        img.width = max_wh
        img.height = max(1, int(round(h * max_wh / w)))
    else:
        img.height = max_wh
        img.width = max(1, int(round(w * max_wh / h)))


def _anchor_image_to_cell(img, col_0based: int, row_0based: int) -> None:
    """将图片锚定为「单格锚点+固定范围」，使图片仅占该单元格、不浮动覆盖它格。"""
    if OneCellAnchor is None or AnchorMarker is None or XDRPositiveSize2D is None or pixels_to_EMU is None:
        img.anchor = get_column_letter(col_0based + 1) + str(row_0based + 1)
        return
    marker = AnchorMarker(col=col_0based, row=row_0based, colOff=0, rowOff=0)
    ext = XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height))
    img.anchor = OneCellAnchor(_from=marker, ext=ext)


def _build_export_xlsx(
    rows: list[dict],
    sheet_title: str = "导出",
    header_labels: list | None = None,
    images_by_row: dict | None = None,
) -> io.BytesIO:
    """根据行数据构建 xlsx：不包含工作表和行号列，使用表头列名，图片限制在单元格内不覆盖表格。"""
    if not Workbook:
        raise HTTPException(status_code=500, detail="请安装 openpyxl: pip install openpyxl")
    max_cols = max(len(r.get("column_data") or []) for r in rows) if rows else 1
    max_cols = max(max_cols, 1)
    if header_labels:
        header = [str(header_labels[i]).strip() if i < len(header_labels) and header_labels[i] else f"列{i + 1}" for i in range(max_cols)]
    else:
        header = [f"列{i + 1}" for i in range(max_cols)]
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    for col, val in enumerate(header, start=1):
        ws.cell(row=1, column=col, value=val)
    # 确定「图片」列并固定列宽，使图片锚定在该格内
    pic_col = 1
    if header_labels:
        for idx, lbl in enumerate(header_labels):
            if (lbl or "").strip() == "图片":
                pic_col = idx
                break
    if get_column_letter:
        ws.column_dimensions[get_column_letter(pic_col + 1)].width = 12
    for row_idx, r in enumerate(rows):
        excel_row = row_idx + 2
        col_data = r.get("column_data") or []
        row_id = r.get("id")
        has_image_in_col = set()
        if images_by_row and row_id and row_id in images_by_row:
            for img_info in images_by_row[row_id]:
                col_ref = img_info.get("col_ref") or ""
                if col_ref.startswith("col_"):
                    try:
                        ci = int(col_ref.replace("col_", ""))
                        if 0 <= ci < max_cols:
                            has_image_in_col.add(ci)
                    except ValueError:
                        pass
        for c in range(max_cols):
            if c in has_image_in_col:
                ws.cell(row=excel_row, column=c + 1, value="")
            else:
                cell_val = col_data[c] if c < len(col_data) else ""
                if cell_val is None:
                    cell_val = ""
                ws.cell(row=excel_row, column=c + 1, value=cell_val)
        if OpenpyxlImage and get_column_letter:
            pic_col_embedded = False
            # 先按 DB 中的图片记录嵌入
            if images_by_row and row_id and row_id in images_by_row:
                for img_info in images_by_row[row_id]:
                    col_ref = img_info.get("col_ref") or ""
                    if not col_ref.startswith("col_"):
                        continue
                    try:
                        ci = int(col_ref.replace("col_", ""))
                    except ValueError:
                        continue
                    if ci < 0 or ci >= max_cols:
                        continue
                    path_val = img_info.get("image_path") or ""
                    row_index = r.get("row_index")
                    full_path = _resolve_image_path(
                        path_val,
                        row_id=row_id,
                        row_index=row_index,
                        export_row_index=row_idx,
                    )
                    if full_path is None:
                        continue
                    try:
                        abs_path = Path(full_path).resolve()
                        img = OpenpyxlImage(str(abs_path))
                        _scale_image_to_cell(img)
                        _anchor_image_to_cell(img, ci, excel_row - 1)
                        ws.add_image(img)
                        if ci == pic_col:
                            pic_col_embedded = True
                        ws.row_dimensions[excel_row].height = max(ws.row_dimensions[excel_row].height or 0, 60)
                    except Exception as e:
                        import sys
                        print(f"[导出图片] 嵌入失败 row={excel_row} col={ci}: {e}", file=sys.stderr)
            # 若该行在「图片」列尚未嵌入，按约定尝试 cellimg_0, cellimg_1...（导出顺序）或 cellimg_{row_id}/{row_index}
            if not pic_col_embedded:
                row_index = r.get("row_index")
                full_path = _resolve_image_path(
                    "",
                    row_id=row_id,
                    row_index=row_index,
                    export_row_index=row_idx,
                )
                if full_path is not None:
                    try:
                        abs_path = Path(full_path).resolve()
                        img = OpenpyxlImage(str(abs_path))
                        _scale_image_to_cell(img)
                        _anchor_image_to_cell(img, pic_col, excel_row - 1)
                        ws.add_image(img)
                        ws.cell(row=excel_row, column=pic_col + 1, value="")
                        ws.row_dimensions[excel_row].height = max(ws.row_dimensions[excel_row].height or 0, 60)
                    except Exception as e:
                        import sys
                        print(f"[导出图片] 嵌入失败 row={excel_row} cellimg: {e}", file=sys.stderr)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.post("/api/export-excel-by-ids")
def api_export_excel_by_ids(body: ExportByIdsRequest):
    """按行 id 导出：传入当前搜索结果的行 id 列表，返回这些行的 Excel 文件（含图片、原表头，无工作表/行号列）。"""
    ids = [int(x) for x in (body.ids or []) if x is not None]
    if not ids:
        raise HTTPException(status_code=400, detail="请先进行搜索，再导出当前结果")
    conn = get_connection()
    try:
        rows = get_full_rows_by_ids(conn, ids)
        header_labels = get_sheet_header_labels(conn, rows[0]["sheet_name"]) if rows else None
        images_by_row = get_images_by_row_ids(conn, ids)
    finally:
        conn.close()
    if not rows:
        raise HTTPException(status_code=404, detail="未找到对应行数据")
    buf = _build_export_xlsx(rows, "搜索结果导出", header_labels=header_labels, images_by_row=images_by_row)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=search_results_export.xlsx"},
    )


# ---------- 前端首页 ----------

@app.get("/")
def index():
    index_file = STATIC_ROOT / "index.html"
    if index_file.exists():
        return FileResponse(index_file)
    return {"message": "请将 index.html 放到 static 目录，或访问 /api/search?q=关键词"}
